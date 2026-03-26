"""
auto_extraccion.py  –  Extracción y análisis colexémico de perífrasis verbales
================================================================================

"""

import argparse
import math
import os
import re
import unicodedata
from collections import Counter

import numpy as np
import pandas as pd

# =========================
# RUTAS
# =========================
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def _rel(*parts):
    return os.path.join(_SCRIPT_DIR, *parts)

_DEFAULT_CORPUS_DIR = _rel("raw_txt_clean")
_DEFAULT_OUT_DIR    = _rel("outputs")
_DEFAULT_HSMS_SRC   = _rel("hsms.src")

# -- argparse ---------------------------------------------------------------
parser = argparse.ArgumentParser(description="Extracción de perífrasis verbales + colexemas")
parser.add_argument("--corpus", default=_DEFAULT_CORPUS_DIR, help="Carpeta de .txt limpios")
parser.add_argument("--outdir", default=_DEFAULT_OUT_DIR,    help="Carpeta de salida")
parser.add_argument("--hsms",   default=_DEFAULT_HSMS_SRC,   help="Fichero hsms.src")
parser.add_argument("--spacy",  default="es_core_news_md",   help="Modelo spaCy")
args = parser.parse_args()

CORPUS_DIR  = args.corpus
OUT_DIR     = args.outdir
HSMS_SRC    = args.hsms
SPACY_MODEL = args.spacy
# ---------------------------------------------------------------------------

if not os.path.isdir(CORPUS_DIR):
    raise FileNotFoundError(
        f"Corpus no exist: {CORPUS_DIR}\n"
        f"Please conform your rute, or use --corpus to indicate a correct rute."
    )

SUPPLEMENT_TXT = os.path.join(OUT_DIR, "supplement_inf_lemmas.txt")
# Add new lemmas to supplement_inf_lemmas.txt manually（p.ej. "custodiar"），
# Next run will be add to STRICT_SET。

NOISE_MAX       = 4
WINDOW_WORDS    = 30
CANDIDATES_TOPN = 500

# collex
MIN_A_DEFAULT = 1
MIN_IN_CORPUS = 5
USE_OR_SMOOTH = True
FDR_PER_PV    = True

# Salida (3 XLSX principales)
OUT_A_XLSX = os.path.join(OUT_DIR, "PV_instances_summary_concordancia.xlsx")
OUT_B_XLSX = os.path.join(OUT_DIR, "Baseline_strict_candidates_patched.xlsx")
OUT_D_XLSX = os.path.join(OUT_DIR, "Collex_results.xlsx")

# CSV auxiliares (for debug)
OUT_BASELINE_STRICT_CSV  = os.path.join(OUT_DIR, "B_baseline_strict.csv")
OUT_BASELINE_CAND_CSV    = os.path.join(OUT_DIR, "B_baseline_candidates.csv")
OUT_BASELINE_PATCHED_CSV = os.path.join(OUT_DIR, "B_baseline_strict_patched.csv")
OUT_MISSING_KEYS_CSV     = os.path.join(OUT_DIR, "B_missing_keys_strict.csv")
OUT_DFA_WITH_C_CSV       = os.path.join(OUT_DIR, "C_dfA_with_c_final.csv")

os.makedirs(OUT_DIR, exist_ok=True)

# =========================
# A) IO + NORMALIZACIÓN + TOKENIZACIÓN
# =========================
def read_txt(fp: str) -> str:
    with open(fp, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def iter_txt_files(folder: str):
    for root, _, files in os.walk(folder):
        for fn in files:
            if fn.lower().endswith(".txt"):
                yield os.path.join(root, fn)

WORD_RE = re.compile(r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ]+")
PUNC_RE = re.compile(r"[.,;:¡!¿?\(\)\[\]\{\}«»\"'—–\-_/\\]+")

def normalize_text(s: str) -> str:
    s = unicodedata.normalize("NFC", s)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"([A-Za-zÁÉÍÓÚÜÑáéíóúüñ]{2,})-\s*\n\s*([A-Za-zÁÉÍÓÚÜÑáéíóúüñ]{2,})", r"\1\2", s)
    s = re.sub(r"([A-Za-zÁÉÍÓÚÜÑáéíóúüñ]{2,})-\s+([a-záéíóúüñ]{2,})", r"\1\2", s)
    s = re.sub(r"\n+", " ", s)
    s = re.sub(r"[ \t]+", " ", s).strip()
    return s

def tokenize_with_spans(s: str):
    tokens = []
    i = 0
    while i < len(s):
        m1 = WORD_RE.match(s, i)
        if m1:
            a, b = m1.start(), m1.end()
            orig = m1.group(0)
            norm = unicodedata.normalize("NFC", orig).lower()
            tokens.append((a, b, orig, norm, "W"))
            i = b
            continue
        m2 = PUNC_RE.match(s, i)
        if m2:
            a, b = m2.start(), m2.end()
            orig = m2.group(0)
            tokens.append((a, b, orig, orig, "P"))
            i = b
            continue
        i += 1
    return tokens

# =========================
# A) Formas verbales auxiliares
# =========================
HABER_FORMS = {
    "he", "has", "ha", "hemos", "habeis", "habéis", "han", "hay",
    "habia", "había", "habias", "habías", "habiamos", "habíamos",
    "habiais", "habíais", "habian", "habían",
    "hube", "hubiste", "hubo", "hubimos", "hubisteis", "hubieron",
    "habra", "habrá", "habras", "habrás", "habremos", "habréis",
    "habreis", "habran", "habrán",
    "habria", "habría", "habrias", "habrías", "habriamos", "habríamos",
    "habriais", "habríais", "habrian", "habrían",
    "haya", "hayas", "hayamos", "hayais", "hayáis", "hayan",
    "hubiera", "hubieras", "hubieramos", "hubiéramos", "hubierais", "hubieran",
    "hubiese", "hubieses", "hubiesemos", "hubiésemos", "hubieseis", "hubiesen",
    "hubiere", "hubieres", "hubieremos", "hubiereis", "hubieren",
    "haber", "habido",
}

TENER_FORMS = {
    "tengo", "tienes", "tiene", "tenemos", "teneis", "tenéis", "tienen",
    "tenia", "tenía", "tenias", "tenías", "teniamos", "teníamos",
    "teniais", "teníais", "tenian", "tenían",
    "tuve", "tuviste", "tuvo", "tuvimos", "tuvisteis", "tuvieron",
    "tendra", "tendrá", "tendras", "tendrás", "tendremos",
    "tendreis", "tendréis", "tendran", "tendrán",
    "tendria", "tendría", "tendrias", "tendrías", "tendriamos", "tendríamos",
    "tendriais", "tendríais", "tendrian", "tendrían",
    "tenga", "tengas", "tengamos", "tengais", "tengáis", "tengan",
    "tuviera", "tuvieras", "tuviéramos", "tuvieramos", "tuvierais", "tuvieran",
    "tuviese", "tuvieses", "tuviésemos", "tuviesemos", "tuvieseis", "tuviesen",
    "tuviere", "tuvieres", "tuviéremos", "tuviereis", "tuvieren",
    "tener", "tenido",
}

DEBER_FORMS = {
    "debo", "debes", "debe", "debemos", "debeis", "debéis", "deben",
    "debia", "debía", "debias", "debías", "debiamos", "debíamos",
    "debiais", "debíais", "debian", "debían",
    "debi", "debí", "debiste", "debio", "debió", "debimos", "debisteis", "debieron",
    "debere", "deberé", "deberas", "deberás", "deberemos",
    "debereis", "deberéis", "deberan", "deberán",
    "deberia", "debería", "deberias", "deberías", "deberiamos", "deberíamos",
    "deberiais", "deberíais", "deberian", "deberían",
    "deba", "debas", "debamos", "debáis", "debais", "deban",
    "debiera", "debieras", "debiéramos", "debieramos", "debierais", "debieran",
    "debiese", "debieses", "debiésemos", "debiesemos", "debieseis", "debiesen",
    "debiere", "debieres", "debiéremos", "debiereis", "debieren",
    "deber", "debido",
}

CLITICS = {"me", "te", "se", "nos", "os", "lo", "la", "los", "las", "le", "les"}
SHORT_INF_WHITELIST = {"ser", "ir", "ver", "dar"}

def split_inf_and_clitics(tok: str):
    base = tok
    tail = ""
    for _ in range(2):
        matched = False
        for c in sorted(CLITICS, key=len, reverse=True):
            if base.endswith(c) and len(base) > len(c):
                tail = c + tail
                base = base[: -len(c)]
                matched = True
                break
        if not matched:
            break
    return base, tail

def is_valid_inf(tok: str) -> bool:
    base, _ = split_inf_and_clitics(tok)
    if base in SHORT_INF_WHITELIST:
        return True
    return base.endswith(("ar", "er", "ir")) and len(base) >= 4

def is_noise(t):
    return t[4] == "P"

def advance_over_noise(tokens, idx, max_noise):
    used = 0
    while idx < len(tokens) and used < max_noise and is_noise(tokens[idx]):
        idx += 1
        used += 1
    return idx

def find_next_word(tokens, idx, max_steps):
    steps = 0
    j = idx
    while j < len(tokens) and steps <= max_steps:
        if tokens[j][4] == "W":
            return j
        j += 1
        steps += 1
    return None

def try_repair_2(tokens, k):
    k2 = find_next_word(tokens, k + 1, max_steps=2)
    if k2 is None:
        return None
    a = tokens[k][3]
    b = tokens[k2][3]
    if b not in {"ar", "er", "ir"}:
        return None
    joined = a + b
    if is_valid_inf(joined):
        base, _ = split_inf_and_clitics(joined)
        return joined, base, tokens[k2][1], k2
    return None

def read_inf_at(tokens, k):
    if k is None or k >= len(tokens) or tokens[k][4] != "W":
        return None
    t1 = tokens[k][3]
    if is_valid_inf(t1):
        base, _ = split_inf_and_clitics(t1)
        return (t1, base, tokens[k][1], k)
    repaired = try_repair_2(tokens, k)
    if repaired:
        return repaired
    return None

STOPWORDS_AFTER_DE = {
    "su", "sus", "mi", "mis", "tu", "tus",
    "nuestro", "nuestra", "nuestros", "nuestras",
    "el", "la", "los", "las", "un", "una", "unos", "unas",
    "este", "esta", "estos", "estas",
    "ese", "esa", "esos", "esas",
    "aquel", "aquella", "aquellos", "aquellas",
    "algún", "alguna", "algunos", "algunas",
    "ningún", "ninguna", "ningunos", "ningunas",
    "otro", "otra", "otros", "otras",
    "todo", "toda", "todos", "todas",
    "cierto", "cierta", "ciertos", "ciertas",
    "vario", "varia", "varios", "varias",
    "mucho", "mucha", "muchos", "muchas",
    "poco", "poca", "pocos", "pocas",
    "demasiado", "demasiada", "demasiados", "demasiadas",
    "bastante", "bastantes",
    "tal", "tales",
    "cada", "cualquier", "cualquiera", "cualesquiera",
}

def extract_pvs_from_text(text: str, filename: str):
    period = "ALL"
    text = normalize_text(text)
    tokens = tokenize_with_spans(text)

    total_word_tokens = sum(1 for t in tokens if t[4] == "W")

    word_idx = [i for i, t in enumerate(tokens) if t[4] == "W"]
    word_pos = {tok_i: pos for pos, tok_i in enumerate(word_idx)}

    rows = []

    def add(pv, aux_orig, inf_full, inf_base, start_char, end_char, i_aux, k_end):
        span_start, span_end = i_aux, k_end
        pv_txt = " ".join(
            tokens[j][2] for j in range(span_start, span_end + 1) if tokens[j][4] == "W"
        ).strip()

        pos_start = word_pos.get(span_start)
        pos_end   = word_pos.get(span_end)
        if pos_end is None:
            left = None
            for wi in reversed(word_idx):
                if wi <= span_end:
                    left = wi
                    break
            pos_end = word_pos.get(left) if left is not None else None

        left_txt = right_txt = ""
        if pos_start is not None:
            left_positions = word_idx[max(0, pos_start - WINDOW_WORDS): pos_start]
            left_txt = " ".join(tokens[j][2] for j in left_positions)
        if pos_end is not None:
            right_positions = word_idx[pos_end + 1: pos_end + 1 + WINDOW_WORDS]
            right_txt = " ".join(tokens[j][2] for j in right_positions)

        rows.append({
            "period":   period,
            "file":     filename,
            "pv":       pv,
            "aux_form": aux_orig,
            "inf_full": inf_full,
            "inf_base": inf_base,
            "start":    start_char,
            "end":      end_char,
            "left_30":  left_txt,
            "pv_span":  pv_txt,
            "right_30": right_txt,
        })

    n = len(tokens)
    for i in range(n):
        a, b, aux_orig, aux_norm, kind = tokens[i]
        if kind != "W":
            continue

        # ---------- DEBER + (DE) + INF ----------
        if aux_norm in DEBER_FORMS:
            j = advance_over_noise(tokens, i + 1, NOISE_MAX)
            if j < n and tokens[j][4] == "W" and tokens[j][3] == "de":
                j = advance_over_noise(tokens, j + 1, NOISE_MAX)
            k = find_next_word(tokens, j, max_steps=NOISE_MAX + 2)
            inf = read_inf_at(tokens, k)
            if inf:
                inf_full, inf_base, end_char, k_end = inf
                add("PV_DEBER_INF", aux_orig, inf_full, inf_base, tokens[i][0], end_char, i, k_end)

        # ---------- TENER QUE + INF ----------
        if aux_norm in TENER_FORMS:
            j = advance_over_noise(tokens, i + 1, NOISE_MAX)
            if j < n and tokens[j][4] == "W" and tokens[j][3] == "que":
                j = advance_over_noise(tokens, j + 1, NOISE_MAX)
                k = find_next_word(tokens, j, max_steps=NOISE_MAX + 2)
                inf = read_inf_at(tokens, k)
                if inf:
                    inf_full, inf_base, end_char, k_end = inf
                    add("PV_TENER_QUE_INF", aux_orig, inf_full, inf_base, tokens[i][0], end_char, i, k_end)

        # ---------- TENER DE + INF ----------
        if aux_norm in TENER_FORMS:
            j = advance_over_noise(tokens, i + 1, NOISE_MAX)
            if j < n and tokens[j][4] == "W" and tokens[j][3] == "de":
                j2 = advance_over_noise(tokens, j + 1, NOISE_MAX)
                k_peek = find_next_word(tokens, j2, max_steps=2)
                if k_peek is not None and tokens[k_peek][3] in STOPWORDS_AFTER_DE:
                    pass  # siguiente token es determinante → no es perífrasis
                else:
                    k = find_next_word(tokens, j2, max_steps=NOISE_MAX + 2)
                    inf = read_inf_at(tokens, k)
                    if inf:
                        inf_full, inf_base, end_char, k_end = inf
                        add("PV_TENER_DE_INF", aux_orig, inf_full, inf_base, tokens[i][0], end_char, i, k_end)

        # ---------- HABER QUE + INF ----------
        if aux_norm in HABER_FORMS:
            j = advance_over_noise(tokens, i + 1, NOISE_MAX)
            if j < n and tokens[j][4] == "W" and tokens[j][3] == "que":
                j = advance_over_noise(tokens, j + 1, NOISE_MAX)
                k = find_next_word(tokens, j, max_steps=NOISE_MAX + 2)
                inf = read_inf_at(tokens, k)
                if inf:
                    inf_full, inf_base, end_char, k_end = inf
                    add("PV_HABER_QUE_INF", aux_orig, inf_full, inf_base, tokens[i][0], end_char, i, k_end)

        # ---------- HABER DE + INF ----------
        if aux_norm in HABER_FORMS:
            j = advance_over_noise(tokens, i + 1, NOISE_MAX)
            if j < n and tokens[j][4] == "W" and tokens[j][3] == "de":
                j = advance_over_noise(tokens, j + 1, NOISE_MAX)
                k = find_next_word(tokens, j, max_steps=NOISE_MAX + 2)
                inf = read_inf_at(tokens, k)
                if inf:
                    inf_full, inf_base, end_char, k_end = inf
                    add("PV_HABER_DE_INF", aux_orig, inf_full, inf_base, tokens[i][0], end_char, i, k_end)

    return rows, total_word_tokens

# =========================
# Helpers de Excel
# =========================
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

def write_df(ws, df: pd.DataFrame, start_row=1, freeze=True):
    header_font = Font(bold=True)
    left = Alignment(vertical="center", horizontal="left")

    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.font = header_font
        cell.alignment = left

    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 42)

    if freeze:
        ws.freeze_panes = ws["A2"]

def apply_number_formats(ws, df: pd.DataFrame):
    fmt_int   = "0"
    fmt_float = "0.0000"
    fmt_sci   = "0.00E+00"

    col_formats = {}
    for idx, col in enumerate(df.columns, start=1):
        c = col.lower()
        if c in {"a", "total_pv", "total_inf", "c_final", "c_spacy",
                 "c_fallback", "tokens", "pv_instances", "files"}:
            col_formats[idx] = fmt_int
        elif c in {"p", "q"}:
            col_formats[idx] = fmt_sci
        elif c in {"oddsratio"}:
            col_formats[idx] = fmt_float
        elif c in {"exp", "score", "neglog10p", "neglog10q"}:
            col_formats[idx] = fmt_float

    for col_idx, numfmt in col_formats.items():
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_idx).number_format = numfmt

# =========================
# BLOQUE A: extracción de PV
# =========================
files = list(iter_txt_files(CORPUS_DIR))
if not files:
    raise FileNotFoundError(
        f"Can't found any .txt document: {CORPUS_DIR}"
    )

all_rows = []
total_tokens = 0

for fp in files:
    txt = read_txt(fp)
    rows, tok_n = extract_pvs_from_text(txt, os.path.basename(fp))
    total_tokens += tok_n
    all_rows.extend(rows)

dfA = pd.DataFrame(all_rows)
cnt = Counter(dfA["pv"].tolist()) if len(dfA) else Counter()

dfSummary = pd.DataFrame([{
    "period":           "ALL",
    "files":            len(files),
    "tokens":           total_tokens,
    "pv_instances":     len(dfA),
    "PV_DEBER_INF":     cnt.get("PV_DEBER_INF", 0),
    "PV_TENER_QUE_INF": cnt.get("PV_TENER_QUE_INF", 0),
    "PV_TENER_DE_INF":  cnt.get("PV_TENER_DE_INF", 0),
    "PV_HABER_QUE_INF": cnt.get("PV_HABER_QUE_INF", 0),
    "PV_HABER_DE_INF":  cnt.get("PV_HABER_DE_INF", 0),
}])

dfConc = dfA[["file", "pv", "left_30", "pv_span", "right_30"]].copy() if len(dfA) else pd.DataFrame(
    columns=["file", "pv", "left_30", "pv_span", "right_30"]
)

wb = Workbook()
ws1 = wb.active; ws1.title = "Instances"
write_df(ws1, dfA)
ws2 = wb.create_sheet("Summary");      write_df(ws2, dfSummary)
ws3 = wb.create_sheet("Concordancia"); write_df(ws3, dfConc)
wb.save(OUT_A_XLSX)
print("✅ Saved A ->", OUT_A_XLSX)

# =========================
# BLOQUE B + C: baseline strict/candidates + parcheo de claves faltantes
# =========================
def norm_key(s: str) -> str:
    s = unicodedata.normalize("NFC", s or "")
    return s.lower().strip()

def build_inf_lemma_set_from_hsms(src_fp: str) -> set:
    inf_lemmas = set()
    bad = 0
    with open(src_fp, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = re.split(r"\s+", line)
            if len(parts) < 3:
                bad += 1
                continue
            _form, lemma, tag = parts[0], parts[1], parts[2]
            if isinstance(tag, str) and tag.startswith("V"):
                lemma_n = norm_key(lemma)
                if lemma_n.endswith(("ar", "er", "ir")):
                    inf_lemmas.add(lemma_n)
    print(f"[hsms] INF lemmas={len(inf_lemmas)} | bad_lines={bad}")
    return inf_lemmas

# Alert if there doesn't exist hsms.src
if os.path.isfile(HSMS_SRC):
    INF_HSMS = build_inf_lemma_set_from_hsms(HSMS_SRC)
else:
    INF_HSMS = set()
    print(
        f"⚠️  ADVERTENCIA: hsms.src no encontrado en {HSMS_SRC}\n"
        f"   El conjunto STRICT_SET estará incompleto. "
        f"Los resultados pueden no ser fiables."
    )

PV_MANUAL = set(dfA["inf_base"].astype(str).apply(norm_key).tolist()) if len(dfA) else set()

SUPP = set()
if os.path.exists(SUPPLEMENT_TXT):
    with open(SUPPLEMENT_TXT, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            x = norm_key(line.strip())
            if x:
                SUPP.add(x)
print(f"[supplement] loaded: {len(SUPP)} lemmas -> {SUPPLEMENT_TXT}")

STRICT_SET = INF_HSMS | PV_MANUAL | SUPP | set(SHORT_INF_WHITELIST)

# -- Carga de spaCy ---------------------------------------------------------
import spacy
try:
    nlp = spacy.load(SPACY_MODEL, disable=["ner", "parser"])
except OSError as e:
    raise RuntimeError(
        f"Modelo spaCy '{SPACY_MODEL}' no encontrado.\n"
        f"Instálalo con:\n  python -m spacy download {SPACY_MODEL}\n"
    ) from e
nlp.max_length = 10_000_000

def yield_chunks_no_word_cut(txt: str, chunk_size: int = 400_000, backtrack: int = 200):
    n = len(txt); i = 0
    while i < n:
        end = min(i + chunk_size, n)
        if end < n:
            j = end
            lower_bound = max(i, end - backtrack)
            while j > lower_bound and not txt[j - 1].isspace():
                j -= 1
            if j > i:
                end = j
        chunk = txt[i:end]
        if chunk:
            yield chunk
        i = end

def strip_clitics_from_surface(v: str, max_strip: int = 2) -> str:
    base = v
    for _ in range(max_strip):
        hit = False
        for c in sorted(CLITICS, key=len, reverse=True):
            if base.endswith(c) and len(base) > len(c):
                base = base[: -len(c)]
                hit = True
                break
        if not hit:
            break
    return base

def looks_like_inf(x: str) -> bool:
    return (x in SHORT_INF_WHITELIST) or (x.endswith(("ar", "er", "ir")) and len(x) >= 4)

def count_baseline_strict_and_candidates(txt: str, chunk_size: int = 400_000):
    txt = normalize_text(txt)
    chunks = list(yield_chunks_no_word_cut(txt, chunk_size=chunk_size))

    c_strict = Counter()
    c_cand   = Counter()

    for doc in nlp.pipe(chunks, batch_size=8):
        for tok in doc:
            if tok.is_space:
                continue
            if tok.pos_ not in {"VERB", "AUX"}:
                continue
            if "Inf" not in tok.morph.get("VerbForm"):
                continue

            lemma = norm_key(tok.lemma_)
            if not looks_like_inf(lemma):
                surf = norm_key(tok.text)
                base = strip_clitics_from_surface(surf, max_strip=3)
                if not looks_like_inf(base):
                    continue
                key = base
            else:
                key = lemma

            if key in STRICT_SET:
                c_strict[key] += 1
            else:
                c_cand[key] += 1

    return c_strict, c_cand

# -- Ejecuta el baseline ----------------------------------------------------
c_strict_all = Counter()
c_cand_all   = Counter()

for fp in iter_txt_files(CORPUS_DIR):
    s_txt = read_txt(fp)
    c_strict, c_cand = count_baseline_strict_and_candidates(s_txt)
    c_strict_all.update(c_strict)
    c_cand_all.update(c_cand)

df_strict = pd.DataFrame(c_strict_all.most_common(), columns=["inf_key", "c_spacy"])
df_cand   = pd.DataFrame(c_cand_all.most_common(),   columns=["inf_key", "c_candidate"])

df_strict.to_csv(OUT_BASELINE_STRICT_CSV, index=False, encoding="utf-8-sig")
df_cand.head(CANDIDATES_TOPN).to_csv(OUT_BASELINE_CAND_CSV, index=False, encoding="utf-8-sig")

print(f"\n[B] strict baseline: types={len(df_strict)} tokens={int(df_strict['c_spacy'].sum())}")
print(
    f"[B] candidates: types={len(df_cand)} "
    f"tokens={int(df_cand['c_candidate'].sum())} "
    f"| saved top{CANDIDATES_TOPN} -> {OUT_BASELINE_CAND_CSV}"
)

# -- C) Parcheo -------------------------------------------------------------
if len(dfA):
    dfA["inf_key"] = dfA["inf_base"].astype(str).apply(norm_key)
    dfA["c_spacy"] = dfA["inf_key"].map(lambda k: int(c_strict_all.get(k, 0))).astype(int)
    missing_keys = sorted(set(dfA.loc[dfA["c_spacy"] == 0, "inf_key"].tolist()))
else:
    dfA["inf_key"] = pd.Series(dtype=str)
    dfA["c_spacy"]  = pd.Series(dtype=int)
    missing_keys = []

pd.DataFrame({"inf_key": missing_keys}).to_csv(OUT_MISSING_KEYS_CSV, index=False, encoding="utf-8-sig")
print(f"[C] strict missing (A but c_spacy=0): {len(missing_keys)} | saved -> {OUT_MISSING_KEYS_CSV}")

missing_set = set(missing_keys)

def count_targets_in_text_exact(txt: str, targets: set) -> Counter:
    txt = normalize_text(txt)
    c = Counter()
    for m in WORD_RE.finditer(txt):
        w    = unicodedata.normalize("NFC", m.group(0)).lower()
        base = strip_clitics_from_surface(w, max_strip=3)
        if base in targets:
            c[base] += 1
    return c

fallback_counter = Counter()
if missing_set:
    for fp in iter_txt_files(CORPUS_DIR):
        fallback_counter.update(count_targets_in_text_exact(read_txt(fp), missing_set))

df_patch = df_strict.copy()
df_patch["c_fallback"] = 0

strict_keys = set(df_patch["inf_key"].tolist())
extra_rows = [
    {"inf_key": k, "c_spacy": 0, "c_fallback": 0}
    for k in missing_keys if k not in strict_keys
]
if extra_rows:
    df_patch = pd.concat([df_patch, pd.DataFrame(extra_rows)], ignore_index=True)

df_patch.loc[df_patch["inf_key"].isin(missing_set), "c_fallback"] = (
    df_patch.loc[df_patch["inf_key"].isin(missing_set), "inf_key"]
    .map(lambda k: int(fallback_counter.get(k, 0)))
    .fillna(0)
    .astype(int)
)

df_patch["c_final"] = df_patch["c_spacy"] + df_patch["c_fallback"]
df_patch.sort_values(["c_final", "c_spacy"], ascending=False, inplace=True)

df_patch.to_csv(OUT_BASELINE_PATCHED_CSV, index=False, encoding="utf-8-sig")
print(f"[C] patched strict baseline saved -> {OUT_BASELINE_PATCHED_CSV}")

c_final_map = dict(zip(df_patch["inf_key"], df_patch["c_final"]))
if len(dfA):
    dfA["c_fallback"] = 0
    dfA.loc[dfA["inf_key"].isin(missing_set), "c_fallback"] = (
        dfA.loc[dfA["inf_key"].isin(missing_set), "inf_key"]
        .map(lambda k: int(fallback_counter.get(k, 0)))
        .fillna(0)
        .astype(int)
    )
    dfA["c_final"] = dfA["inf_key"].map(lambda k: int(c_final_map.get(k, 0))).astype(int)

dfA.to_csv(OUT_DFA_WITH_C_CSV, index=False, encoding="utf-8-sig")
print(f"[C] dfA with c_spacy/c_fallback/c_final -> {OUT_DFA_WITH_C_CSV}")

# Guarda B → XLSX
wb = Workbook()
ws1 = wb.active; ws1.title = "Strict_patched"
write_df(ws1, df_patch); apply_number_formats(ws1, df_patch)

ws2 = wb.create_sheet("Candidates_top")
write_df(ws2, df_cand.head(CANDIDATES_TOPN))

ws3 = wb.create_sheet("Missing_strict")
write_df(ws3, pd.DataFrame({"inf_key": missing_keys}))

ws4 = wb.create_sheet("Strict_raw_spacy")
write_df(ws4, df_strict); apply_number_formats(ws4, df_strict)

wb.save(OUT_B_XLSX)
print("✅ Saved B ->", OUT_B_XLSX)

# =========================
# BLOQUE D: Análisis colexémico (Fisher + BH-FDR + EXP)
# =========================
try:
    from scipy.stats import fisher_exact as scipy_fisher_exact
except ImportError:
    raise RuntimeError("scipy no está instalado. Ejecuta: pip install scipy")

def fisher_exact_2x2(a, b, c, d):
    OR, p = scipy_fisher_exact([[a, b], [c, d]], alternative="two-sided")
    return float(OR), float(p)

def or_smooth(a, b, c, d):
    return ((a + 0.5) * (d + 0.5)) / ((b + 0.5) * (c + 0.5))

def signed_logp(p, oddsratio):
    p    = max(float(p), 1e-300)
    sign = 1 if oddsratio > 1 else (-1 if oddsratio < 1 else 0)
    return sign * (-math.log10(p))

def bh_fdr(pvals):
    pvals = np.asarray(pvals, dtype=float)
    n = len(pvals)
    if n == 0:
        return np.array([])
    order  = np.argsort(pvals)
    ranked = pvals[order]
    q = ranked * n / (np.arange(1, n + 1))
    q = np.minimum.accumulate(q[::-1])[::-1]
    q = np.clip(q, 0, 1)
    out        = np.empty(n, dtype=float)
    out[order] = q
    return out

dfC = pd.DataFrame()

if len(dfA):
    base_u    = dfA[["inf_key", "c_final"]].drop_duplicates()
    TOTAL_INF = int(base_u["c_final"].sum())

    pv_totals = dfA.groupby("pv").size().rename("total_PV").reset_index()
    pv_inf = (
        dfA.groupby(["pv", "inf_key"], as_index=False)
           .size()
           .rename(columns={"size": "a"})
    ).merge(base_u, on="inf_key", how="left").merge(pv_totals, on="pv", how="left")

    inf_label = (
        dfA.groupby("inf_key")["inf_base"]
           .agg(lambda s: s.value_counts().index[0])
           .rename("INF")
           .reset_index()
    )
    pv_inf = pv_inf.merge(inf_label, on="inf_key", how="left")

    rows_d = []
    for _, r in pv_inf.iterrows():
        pv       = r["pv"]
        a        = int(r["a"])
        Fv       = int(r["c_final"])
        total_PV = int(r["total_PV"])

        if a < MIN_A_DEFAULT:
            continue
        if Fv < MIN_IN_CORPUS:
            continue
        if a > Fv:
            continue

        b = total_PV - a
        c = Fv - a
        d = (TOTAL_INF - total_PV) - c
        if min(b, c, d) < 0:
            continue

        OR_raw, p = fisher_exact_2x2(a, b, c, d)
        OR_use    = or_smooth(a, b, c, d) if USE_OR_SMOOTH else OR_raw
        score     = signed_logp(p, OR_use)
        EXP       = total_PV * (Fv / TOTAL_INF)

        rows_d.append({
            "pv":        pv,
            "INF":       r["INF"],
            "inf_key":   r["inf_key"],
            "a":         a,
            "exp":       float(EXP),
            "c_final":   Fv,
            "total_PV":  total_PV,
            "total_INF": TOTAL_INF,
            "oddsratio": OR_use,
            "p":         p,
            "score":     score,
        })

    dfC = pd.DataFrame(rows_d)

if len(dfC):
    if FDR_PER_PV:
        dfC["q"] = 0.0
        for pv, idx in dfC.groupby("pv").groups.items():
            dfC.loc[idx, "q"] = bh_fdr(dfC.loc[idx, "p"].values)
    else:
        dfC["q"] = bh_fdr(dfC["p"].values)

    dfC["direction"] = dfC["oddsratio"].apply(
        lambda x: "atraido" if x > 1 else ("repelido" if x < 1 else "neutral")
    )
    dfC["neglog10p"] = dfC["p"].apply(lambda x: -math.log10(max(float(x), 1e-300)))
    dfC["neglog10q"] = dfC["q"].apply(lambda x: -math.log10(max(float(x), 1e-300)))
    dfC = dfC.sort_values(["pv", "score"], ascending=[True, False]).reset_index(drop=True)

df_sig = dfC[dfC["q"] < 0.05].copy() if len(dfC) else pd.DataFrame()

# Guarda D → XLSX
wb = Workbook()
ws1 = wb.active; ws1.title = "ALL"
write_df(ws1, dfC); apply_number_formats(ws1, dfC)
ws2 = wb.create_sheet("q<0.05")
write_df(ws2, df_sig); apply_number_formats(ws2, df_sig)
wb.save(OUT_D_XLSX)
print("✅ Saved D ->", OUT_D_XLSX)

print("\nDONE ✅  Salida:")
print(" -", OUT_A_XLSX)
print(" -", OUT_B_XLSX)
print(" -", OUT_D_XLSX)
print(
    f"\nConsejo: revisa los candidatos en B XLSX (hoja 'Candidates_top').\n"
    f"Si encuentras variantes históricas reales, añádelas una por línea a:\n"
    f"  {SUPPLEMENT_TXT}\n"
    f"y vuelve a ejecutar el script."
)
